#!/usr/bin/env python3
"""Convert victron-modbus-json's JSON output into an Excel workbook.

One sheet per discovered device (unit ID + service type), plus a Summary
sheet listing every device and an Overview sheet with the scan metadata.

Usage:
    python3 json_to_excel.py victron.json [-o victron.xlsx]
"""
import argparse
import json
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill(start_color="1F4E5F", end_color="1F4E5F", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
UNAVAILABLE_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
NOTE_FILL = PatternFill(start_color="FDEBD0", end_color="FDEBD0", fill_type="solid")

REGISTER_COLUMNS = [
    ("path", "Path"),
    ("register", "Register"),
    ("dataType", "Data Type"),
    ("access", "Access"),
    ("unit", "Unit"),
    ("raw", "Raw"),
    ("value", "Value"),
    ("label", "Label"),
    ("available", "Available"),
    ("note", "Note"),
]


def cell_value(raw):
    """openpyxl can't store lists/dicts directly; flatten those to text."""
    if isinstance(raw, (list, dict)):
        return json.dumps(raw)
    return raw


def sheet_name_for(unit_id, service_type, used_names):
    short = service_type.replace("com.victronenergy.", "")
    base = f"{unit_id}_{short}"[:31]
    name = base
    suffix = 2
    while name in used_names:
        name = f"{base[: 31 - len(str(suffix)) - 1]}_{suffix}"
        suffix += 1
    used_names.add(name)
    return name


def autosize_columns(ws, min_width=8, max_width=60):
    widths = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            length = len(str(cell.value))
            col = cell.column_letter
            widths[col] = max(widths.get(col, min_width), min(length + 2, max_width))
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def style_header(ws, ncols, row=1):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def build_overview_sheet(wb, data):
    ws = wb.active
    ws.title = "Overview"
    rows = [
        ("Host", data.get("host")),
        ("Port", data.get("port")),
        ("Scanned At", data.get("scannedAt")),
        ("Unit ID Range", f"{data.get('unitIdRange', ['', ''])[0]} - {data.get('unitIdRange', ['', ''])[1]}"),
        ("Devices Found", len(data.get("devices", []))),
    ]
    for r, (label, value) in enumerate(rows, start=1):
        ws.cell(row=r, column=1, value=label).font = Font(bold=True)
        ws.cell(row=r, column=2, value=value)

    warnings = data.get("warnings") or []
    if warnings:
        start = len(rows) + 2
        ws.cell(row=start, column=1, value="Warnings").font = Font(bold=True)
        for i, w in enumerate(warnings):
            ws.cell(row=start + 1 + i, column=1, value=w)
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 60


def build_summary_sheet(wb, devices, sheet_names):
    ws = wb.create_sheet("Summary")
    headers = ["Unit ID", "Device Instance", "Service Type", "Register Count", "Available", "Sheet"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h)
    for r, (dev, sheet_name) in enumerate(zip(devices, sheet_names), start=2):
        registers = dev.get("registers", [])
        available_count = sum(1 for reg in registers if reg.get("available", True) and reg.get("value") is not None)
        ws.cell(row=r, column=1, value=dev.get("unitId"))
        ws.cell(row=r, column=2, value=dev.get("deviceInstance"))
        ws.cell(row=r, column=3, value=dev.get("serviceType"))
        ws.cell(row=r, column=4, value=len(registers))
        ws.cell(row=r, column=5, value=available_count)
        cell = ws.cell(row=r, column=6, value=sheet_name)
        cell.hyperlink = f"#'{sheet_name}'!A1"
        cell.font = Font(color="0000FF", underline="single")
    style_header(ws, len(headers))
    autosize_columns(ws)


def build_device_sheet(wb, dev, sheet_name):
    ws = wb.create_sheet(sheet_name)
    for c, (_, label) in enumerate(REGISTER_COLUMNS, start=1):
        ws.cell(row=1, column=c, value=label)
    for r, reg in enumerate(dev.get("registers", []), start=2):
        for c, (key, _) in enumerate(REGISTER_COLUMNS, start=1):
            ws.cell(row=r, column=c, value=cell_value(reg.get(key)))
        if reg.get("note"):
            fill = NOTE_FILL
        elif not reg.get("available", True):
            fill = UNAVAILABLE_FILL
        else:
            fill = None
        if fill:
            for c in range(1, len(REGISTER_COLUMNS) + 1):
                ws.cell(row=r, column=c).fill = fill
    style_header(ws, len(REGISTER_COLUMNS))
    autosize_columns(ws)


def convert(json_path, xlsx_path):
    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    devices = data.get("devices", [])
    wb = Workbook()
    build_overview_sheet(wb, data)

    used_names = {"Overview", "Summary"}
    sheet_names = [sheet_name_for(d.get("unitId"), d.get("serviceType", ""), used_names) for d in devices]

    build_summary_sheet(wb, devices, sheet_names)
    for dev, sheet_name in zip(devices, sheet_names):
        build_device_sheet(wb, dev, sheet_name)

    wb.save(xlsx_path)
    return len(devices)


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("json_file", help="path to the JSON produced by victron-modbus-json")
    parser.add_argument("-o", "--out", help="output .xlsx path (default: same name as input, .xlsx extension)")
    args = parser.parse_args()

    out_path = args.out or (args.json_file.rsplit(".", 1)[0] + ".xlsx")
    n = convert(args.json_file, out_path)
    print(f"Wrote {out_path} ({n} device sheet(s))")


if __name__ == "__main__":
    sys.exit(main())
